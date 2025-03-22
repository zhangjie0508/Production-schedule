"""可以对每个表单进行组1分组，每个表的组2按照组1来排序  √
处理逾期交付的订单
逾期交付拆分前面无交期的订单√
把拆分订单移到所有订单最后，不改变其他订单排序，只修改拆分订单的排序√
重新判断是否换料√
重新计算生产开始时间和结束时间√
重新判断是否逾期√
如何循环这些，直到逾期订单前没有可拆分的订单为止
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import re
# 设定文件路径
input_file = r"D:\3.21订单信息.xlsx"
output_file = r"D:\3.21优化排产.xlsx"


# **📌 读取 Excel 数据**
df_original = pd.read_excel(input_file, dtype={"预计交期": str})  # 先保留原始数据
df = df_original.copy()  # 复制用于后续处理

# **📌 先从原始数据里筛选‘差异化’订单 和 '已完成' 订单**
df_other = df_original[
    df_original["加工工艺"].astype(str).str.contains("差异化", na=False, regex=False) |
    df_original["完成量"].astype(str).str.contains("已完成", na=False, regex=False)].copy()

# **📌 处理其他订单（不影响 df_other）**
df["加工工艺"] = df["加工工艺"].astype(str).str.strip().str.replace(r"\s+", "", regex=True)

# **📌 确保‘其他’表单格式与输入表格一致**
df_other = df_other.reindex(columns=df_original.columns)  # 保持列顺序一致

# **📌 处理“来料304不锈钢” 等前缀问题**
def normalize_material(material):
    """ 去掉 '来料' 前缀，确保数据处理时材质一致 """
    return material.replace("来料", "").strip() if isinstance(material, str) else material

def restore_material(original_material, processed_material):
    """ 如果原始材质带有 '来料'，则输出时加回去 """
    return f"来料{processed_material}" if original_material.startswith("来料") else processed_material

# **📌 处理数据时去掉前缀**
df["原始材料材质"] = df["材料材质"]  # 先保存原始数据
df["材料材质"] = df["材料材质"].apply(normalize_material)

# 初始化设备列
df["设备"] = ""

# 计算负荷
yixing1_load, yixing2_load = 0, 0
assigned_materials = {}

# 过滤出非直管、非差异化订单
df_orders = df[~df["加工工艺"].isin(["直管", "差异化"])].copy()

# 确保数据类型正确
df_orders["材料厚度"] = pd.to_numeric(df_orders["材料厚度"], errors="coerce")
df_orders["未完成数量"] = pd.to_numeric(df_orders["未完成数量"], errors="coerce")
df_orders["材料材质"] = df_orders["材料材质"].astype(str)

# **🔹 先按“材料厚度”降序排列，确保厚的订单优先分配**
df_orders = df_orders.sort_values(by="材料厚度", ascending=False)

# 遍历订单，分配设备
for index, row in df_orders.iterrows():
    thickness = row["材料厚度"]
    material = row["材料材质"]
    unfinished = row["未完成数量"] if not pd.isna(row["未完成数量"]) else 0
    load_yixing1, load_yixing2 = unfinished / 50, unfinished / 80
    key = (thickness, material)

    # **🔹 规则 1：判断是否必须给“异型管机2”**
    if thickness not in {0.5, 0.75, 0.6, 0.8, 1.0} or "不锈钢" in material:
        df.at[index, "设备"] = "异型管机2"
        yixing2_load += load_yixing2
        assigned_materials[key] = "异型管机2"

    # **🔹 规则 2：如果相同厚度 & 材质的订单已分配过，则沿用原设备**
    elif key in assigned_materials:
        assigned_device = assigned_materials[key]
        df.at[index, "设备"] = assigned_device
        if assigned_device == "异型管机1":
            yixing1_load += load_yixing1
        else:
            yixing2_load += load_yixing2

    # **🔹 规则 3：厚的订单优先分配给“异型管机2”**
    else:
        # **优先选择“异型管机2”**
        if thickness >= 1.0 or yixing2_load <= yixing1_load:
            assigned_device = "异型管机2"
        else:
            assigned_device = "异型管机1"

        df.at[index, "设备"] = assigned_device
        assigned_materials[key] = assigned_device
        if assigned_device == "异型管机1":
            yixing1_load += load_yixing1
        else:
            yixing2_load += load_yixing2

# **🔹 处理“直管机”订单**
df.loc[df["加工工艺"] == "直管", "设备"] = "直管机"

# 清理无效设备数据，去除空设备
df = df[df["设备"] != ""]
df = df.dropna(subset=["设备"])

# 设备结束时间字典
device_last_end_time = {}  # 仅存有效设备


def convert_due_date(due_date):
    if pd.isna(due_date) or due_date.strip() == "":
        return pd.NaT
    due_date = due_date.strip()  # 清除两端空格
    try:
        parts = re.split(r'\s+', due_date)  # 避免多个空格导致 split 出错
        if len(parts) == 2:
            date_part, time_part = parts
            month, day = map(int, date_part.split("."))
            time_part = time_part.replace("：", ":")  # 修正中文冒号
            base_date = f"2025-{month:02d}-{day:02d} {time_part}"
            return pd.to_datetime(base_date, errors="coerce")  # 防止异常日期
    except Exception as e:
        print(f"⚠️ 解析错误: {due_date}，错误信息: {e}")
    return pd.NaT  # 解析失败的也先设为 NaT

# 📌 交期转换
df["交期排序"] = df["预计交期"].apply(convert_due_date)
# ✅ **填充 NaT 为 2100-01-01，而不是最大日期**
df["交期排序"] = df["交期排序"].fillna(pd.Timestamp("2100-01-01"))
# 📌 标记是否有交期 (1: 有交期, 0: 无交期)
df["是否有交期"] = (df["交期排序"] < pd.Timestamp("2100-01-01")).astype(int)

#df["组1"] = df.groupby(["设备","材料厚度", "材料材质"]).ngroup()
# 创建一个新的列 '组1' 用于标识设备内的分组编号
df["组1"] = -1  # 初始化组1列

# 对每个设备子集单独进行分组并排序
for device in df["设备"].unique():
    # 对每个设备的子集进行分组排序
    device_df = df[df["设备"] == device].copy()  # 使用 .copy() 创建一个副本
    device_df["组1"] = device_df.groupby(["材料厚度", "材料材质"]).ngroup()  # 在设备内进行分组
    df.loc[df["设备"] == device, "组1"] = device_df["组1"]  # 将计算出的组1赋值回原 DataFrame



# 📌 计算组的最早交期
# 计算组最早交期时，排除无交期的订单
df["组最早交期"] = df.loc[df["交期排序"] <= pd.Timestamp("2100-01-01"), "交期排序"].groupby(df["组1"]).transform("min")
# ✅ 确保 `组最早交期` 是 datetime 格式
df["组最早交期"] = pd.to_datetime(df["组最早交期"])

# 📌 排序：保证相同材质 & 厚度的订单在一起，同时组外按交期排序
df.sort_values(
    by=["组最早交期", "材料材质", "材料厚度", "是否有交期", "交期排序"],
    ascending=[True, True, True, False, True],
    inplace=True
)

# 是否换料标记
df["是否换料"] = (
    df["设备"].ne(df["设备"].shift()) |
    df["材料厚度"].ne(df["材料厚度"].shift()) |
    df["材料材质"].ne(df["材料材质"].shift())
).map({True: "是", False: "否"})


# **📌 仅调整“异型管机2” 的排序**
df_yixing2 = df[df["设备"] == "异型管机2"].copy()

# **📌 判断是否是“异型管机1”无法生产的订单**
df_yixing2["异型管机1不可生产"] = df_yixing2.apply(
    lambda row: row["材料厚度"] not in [0.5, 0.75, 0.6, 0.8, 1.0] or "不锈钢" in row["材料材质"],
    axis=1
)

# **📌 重新排序“异型管机2” 的订单**
df_yixing2 = df_yixing2.sort_values(
    by=["异型管机1不可生产","组最早交期", "材料材质", "材料厚度", "是否有交期", "交期排序"],
    ascending=[False,True, True, True, False, True]
)

# **📌 重新合并回 df**
df = df[df["设备"] != "异型管机2"]  # 先移除原来的“异型管机2”数据
df = pd.concat([df, df_yixing2], ignore_index=True)  # 重新合并排序后的数据

# **📌 更新“是否换料”**
df["是否换料"] = (df["设备"].ne(df["设备"].shift()) |
                  df["材料厚度"].ne(df["材料厚度"].shift()) |
                  df["材料材质"].ne(df["材料材质"].shift())).map({True: "是", False: "否"})

#对组2进行排序
for device in df["设备"].unique():
    # 过滤当前设备的数据
    device_df = df[df["设备"] == device].copy()  # 使用 .copy() 避免修改原 DataFrame

    # 在当前设备内，基于 材料厚度 和 材料材质 进行分组，生成 组1
    device_df["组1"] = device_df.groupby(["材料厚度", "材料材质"]).ngroup()

    # 生成 组2，确保组1相同的值对应相同的 组2，且从 0 开始编号
    device_df["组2"] = pd.factorize(device_df["组1"])[0]  # 设备内部 factorize

    # 赋值回原 DataFrame
    df.loc[df["设备"] == device, ["组1", "组2"]] = device_df[["组1", "组2"]]




# **📌 计算生产时间**
"""
   根据设备类型计算生产所需时间。
   :param row: DataFrame中的一行数据
   :return: 生产时间（格式：'X小时 Y分钟'）
   """
def calculate_production_time(row):
    if row["设备"] == "直管机":
        hours = row["生产件数"] / 90# 直管机每小时生产90件
    elif row["设备"] == "异型管机1":
        hours = row["未完成数量"] / 50# 异型管机1每小时生产50件
    elif row["设备"] == "异型管机2":
        hours = row["未完成数量"] / 80 # 异型管机2每小时生产80件
    else:
        return "0小时 0分钟"
    total_minutes = round(hours * 60)# 转换为分钟并四舍五入
    return f"{total_minutes // 60}小时 {total_minutes % 60}分钟"

# **📌 计算生产时间**
df["生产时间"] = df.apply(calculate_production_time, axis=1)

# **📌 初始化生产开始时间和结束时间**
df["生产开始时间"] = pd.NaT
df["生产结束时间"] = pd.NaT

# 工作时间段
work_shifts = [
    ("08:00", "12:00"),
    ("13:30", "17:30"),
    ("18:00", "21:00")
]

# 休息时间段
break_times = [
    ("12:00", "13:30"),
    ("17:30", "18:00"),
    ("21:00", "08:00")  # 跨天休息
]


# **📌 获取当前时间所在的班次信息**
def get_next_available_shift(current_time):
    """
    获取当前时间最近的可用工作班次及剩余可用时间。
    :param current_time: 当前时间戳
    :return: (班次开始时间, 班次结束时间, 剩余可用分钟数)
    """
    current_day = current_time.date() #获取当前日期，这用于确保计算的时间点都在同一天，方便格式化时间戳。

    for start, end in work_shifts: #遍历工作时间段（work_shifts），代码会遍历 work_shifts，找到下一个可用的班次。
        shift_start = pd.Timestamp(f"{current_day} {start}") #构造班次时间，将字符串时间转换为 Timestamp，形成完整的时间段。
        shift_end = pd.Timestamp(f"{current_day} {end}")
        # 处理跨天班次，如果 shift_end 是 08:00，说明是跨天班次，需要加一天。
        if shift_end.hour == 8:
            shift_end += pd.Timedelta(days=1)

        if current_time < shift_end:
            """计算当前可用时间，max(current_time, shift_start) 确保当前时间不会早于班次开始时间。
            (shift_end - max(current_time, shift_start)).seconds // 60计算当前时间到班次结束的分钟数。"""
            available_minutes = max(0, (shift_end - max(current_time, shift_start)).seconds // 60)
            return shift_start, shift_end, available_minutes #返回班次信息，找到可用班次后，返回起始时间、结束时间以及剩余可用分钟数。

    # 如果当前时间超出所有班次，进入下一天的第一个班次，如果当前时间已经超出当天所有班次，返回下一天的第一个班次（08:00 - 12:00，共 240 分钟）。
    next_day = current_day + pd.Timedelta(days=1)
    return pd.Timestamp(f"{next_day} 08:00"), pd.Timestamp(f"{next_day} 12:00"), 240


# **📌 处理休息时间**
def is_in_break_time(time):
    """
    判断当前时间是否处于休息时间。
    :param time: 时间戳
    :return: 若在休息时间内，返回休息结束时间；否则返回None
    """
    current_day = time.date()
    for start, end in break_times:
        break_start = pd.Timestamp(f"{current_day} {start}")
        break_end = pd.Timestamp(f"{current_day} {end}")
        #处理跨天休息时间
        if break_end.hour == 8:
            break_end += pd.Timedelta(days=1)  # 处理跨天情况
        if break_start <= time < break_end:
            return break_end
    return None

# **📌 初始化各设备的上次结束时间**
device_last_end_time = {
    "直管机": pd.Timestamp("2025-03-21 08:00"),
    "异型管机1": pd.Timestamp("2025-03-21 08:00"),
    "异型管机2": pd.Timestamp("2025-03-21 08:00")
}

# **📌 计算生产开始时间 & 结束时间**
for index, row in df.iterrows():
    device = row["设备"]
    start_time = device_last_end_time[device] # 获取设备上次结束时间作为新任务开始时间

    # **📌 如果需要换料，增加 15 分钟**
    if row["是否换料"] == "是":
        start_time += pd.Timedelta(minutes=15)
    # 解析生产时间
    total_minutes = int(row["生产时间"].split("小时")[0]) * 60 + int(
        row["生产时间"].split("小时")[1].replace("分钟", ""))
    remaining_time = total_minutes
    segments = []

    while remaining_time > 0:
        # **📌 检查是否在休息时间**
        break_end = is_in_break_time(start_time)
        if break_end:
            start_time = break_end  # 跳过休息时间

        # **📌 获取当前班次**
        shift_start, shift_end, available_minutes = get_next_available_shift(start_time)
        production_time = min(remaining_time, available_minutes) # 计算当前可用生产时间
        end_time = start_time + pd.Timedelta(minutes=production_time)

        segments.append((start_time, end_time))
        remaining_time -= production_time
        start_time = end_time + pd.Timedelta(minutes=1)  # 进入下一个时间段

    # **📌 记录计算结果**
    df.at[index, "生产开始时间"] = segments[0][0]
    df.at[index, "生产结束时间"] = segments[-1][1]
    device_last_end_time[device] = segments[-1][1]  # 更新设备的上次结束时间

# **📌 计算项目交付时间**
df["项目交付时间"] = df.groupby("订单编号")["生产结束时间"].transform("max")

# **📌 生成项目交付时间表**
project_delivery_df = df[["订单编号", "项目交付时间"]].drop_duplicates().sort_values(by="项目交付时间")



#判断是否逾期
# **📌 转换预计交期**
df["预计交期"] = df["预计交期"].apply(convert_due_date)

# **📌 确保 '生产结束时间' 也是 datetime 类型**
df["生产结束时间"] = pd.to_datetime(df["生产结束时间"], errors="coerce")

# **📌 计算是否按时交付**
df["按时交付检查"] = df["预计交期"].isna() | (df["预计交期"] >= df["生产结束时间"])

# **📌 转换 True/False 为 '按时交付' / '逾期交付'**
df["按时交付检查"] = df["按时交付检查"].map({True: "按时交付", False: "逾期交付"})



#判断是否拆分，拆分逾期订单前无交期的订单
def debug_move_orders(df_group):
    """ 在单个 '设备' 子表内计算 '是否拆分' """
    df_copy = df_group.copy()
    df_copy['按时交付检查'] = df_copy['按时交付检查'].fillna('').astype(str).str.strip()
    df_copy['组2'] = pd.to_numeric(df_copy['组2'], errors='coerce')
    df_copy['是否拆分'] = '否'  # 默认值

    overdue_orders = df_copy[df_copy['按时交付检查'] == '逾期交付']
    split_orders = set()

    for index, row in overdue_orders.iterrows():
        group2_id = row['组2']
        prev_group2_id = group2_id - 1
        prev_group_orders = pd.DataFrame()  # ✅ 先初始化为空 DataFrame，防止未定义错误

        while prev_group2_id >= 0:  # 继续往前找，直到组2 == 0
            prev_group_orders = df_copy[df_copy['组2'] == prev_group2_id]  # 每次循环更新

            if prev_group_orders.empty:  # 如果当前组是空的，继续往前找
                prev_group2_id -= 1
                continue

            fixed_date_orders = prev_group_orders[prev_group_orders['交期排序'] == pd.Timestamp("2100-01-01")]

            if not fixed_date_orders.empty:  # 如果找到符合条件的订单，就停止循环
                split_orders.update(fixed_date_orders.index.tolist())
                break  # 找到了就不继续查找

            prev_group2_id -= 1  # 否则继续往前找

        if prev_group_orders.empty:
            continue

        # 确保 "交期排序" 为时间格式
        df_copy['交期排序'] = pd.to_datetime(df_copy['交期排序'], errors='coerce')

        # 过滤 交期排序=2100-01-01 的订单
        fixed_date_orders = prev_group_orders[prev_group_orders['交期排序'] == pd.Timestamp("2100-01-01")]

        split_orders.update(fixed_date_orders.index.tolist())

    df_copy.loc[df_copy.index.isin(split_orders), '是否拆分'] = '是'
    # ✅ 拆分为两个部分：
    df_not_split = df_copy[df_copy['是否拆分'] == '否']  # ❌ 不需要拆分的订单（保持原顺序）
    df_split = df_copy[df_copy['是否拆分'] == '是'].sort_values(
        by=['材料材质', '材料厚度'], ascending=[True, True]  # ✅ 需要拆分的订单（材质+厚度排序）
    )
    # ✅ 按原索引顺序合并
    df_copy = pd.concat([df_not_split, df_split])


    return df_copy  # 确保返回 df_copy

# ✅ 运行 groupby.apply()
df = df.groupby("设备", group_keys=False).apply(debug_move_orders).reset_index(drop=True)
print(df.head())  # 检查 "是否拆分"

# **📌 先按设备分组，确保组内排序**
def sort_by_due_status(df_group):
    """在每个 设备 + 组2 内部，把 '逾期交付' 的订单排在最前面"""
    return df_group.sort_values(by=["组2", "按时交付检查"], ascending=[True, False])

df = df.groupby("设备", group_keys=False).apply(sort_by_due_status).reset_index(drop=True)

#重新判断是否需要换料
# 只修改子表 "直管机" 的数据
zhiguan1 = df["设备"] == "直管机"
df.loc[zhiguan1, "是否换料"] = (
    df.loc[zhiguan1, "设备"].ne(df.loc[zhiguan1, "设备"].shift()) |
    df.loc[zhiguan1, "材料厚度"].ne(df.loc[zhiguan1, "材料厚度"].shift()) |
    df.loc[zhiguan1, "材料材质"].ne(df.loc[zhiguan1, "材料材质"].shift())
).map({True: "是", False: "否"})
# 只修改子表 "异型管1" 的数据
yixing1 = df["设备"] == "异型管机1"
df.loc[yixing1, "是否换料"] = (
    df.loc[yixing1, "设备"].ne(df.loc[yixing1, "设备"].shift()) |
    df.loc[yixing1, "材料厚度"].ne(df.loc[yixing1, "材料厚度"].shift()) |
    df.loc[yixing1, "材料材质"].ne(df.loc[yixing1, "材料材质"].shift())
).map({True: "是", False: "否"})
# 只修改子表 "异型管机2" 的数据
yixing2 = df["设备"] == "异型管机2"
df.loc[yixing2, "是否换料"] = (
    df.loc[yixing2, "设备"].ne(df.loc[yixing2, "设备"].shift()) |
    df.loc[yixing2, "材料厚度"].ne(df.loc[yixing2, "材料厚度"].shift()) |
    df.loc[yixing2, "材料材质"].ne(df.loc[yixing2, "材料材质"].shift())
).map({True: "是", False: "否"})



#重新计算生产开始时间和结束时间
# **📌 重新初始化设备上次结束时间** 每次这里要修改成对应的设备开始时间
device_last_end_time = {device: pd.Timestamp("2025-03-21 8:00") for device in df["设备"].unique()}

# **📌 计算生产开始时间 & 结束时间**
for index, row in df.iterrows():
    device = row["设备"]
    start_time = device_last_end_time[device]  # 获取设备上次结束时间作为新任务开始时间

    # **📌 如果需要换料，增加 15 分钟**
    if row["是否换料"] == "是":
        start_time += pd.Timedelta(minutes=15)

    # **📌 解析生产时间**
    production_str = row["生产时间"]
    if "小时" in production_str and "分钟" in production_str:
        hours, minutes = map(int, production_str.replace("小时", "").replace("分钟", "").split())
    elif "小时" in production_str:
        hours = int(production_str.replace("小时", ""))
        minutes = 0
    elif "分钟" in production_str:
        hours = 0
        minutes = int(production_str.replace("分钟", ""))
    else:
        hours = 0
        minutes = 0

    total_minutes = hours * 60 + minutes
    remaining_time = total_minutes
    segments = []

    while remaining_time > 0:
        # **📌 检查是否在休息时间**
        break_end = is_in_break_time(start_time)
        if break_end:
            start_time = break_end  # 跳过休息时间

        # **📌 获取当前班次**
        shift_start, shift_end, available_minutes = get_next_available_shift(start_time)
        production_time = min(remaining_time, available_minutes)  # 计算当前可用生产时间
        end_time = start_time + pd.Timedelta(minutes=production_time)

        segments.append((start_time, end_time))
        remaining_time -= production_time
        start_time = end_time + pd.Timedelta(minutes=1)  # 进入下一个时间段

    # **📌 记录计算结果**
    df.at[index, "生产开始时间"] = segments[0][0]
    df.at[index, "生产结束时间"] = segments[-1][1]
    device_last_end_time[device] = segments[-1][1]  # 更新设备的上次结束时间

#重新判断是否逾期

# **📌 确保 '生产结束时间' 也是 datetime 类型**
df["生产结束时间"] = pd.to_datetime(df["生产结束时间"], errors="coerce")

# **📌 计算是否按时交付**
df["按时交付检查"] = df["预计交期"].isna() | (df["预计交期"] >= df["生产结束时间"])

# **📌 转换 True/False 为 '按时交付' / '逾期交付'**
df["按时交付检查"] = df["按时交付检查"].map({True: "按时交付", False: "逾期交付"})

#重新计算项目交付时间
# **📌 计算项目交付时间**
df["项目交付时间"] = df.groupby("订单编号")["生产结束时间"].transform("max")

# **📌 生成项目交付时间表**
project_delivery_df = df[["订单编号", "项目交付时间"]].drop_duplicates().sort_values(by="项目交付时间")

# **📌 格式化所有表单的日期**
for table in [df]:
    table["下单日期"] = pd.to_datetime(table["下单日期"]).dt.strftime("%Y-%m-%d").astype(str)
    table["预计交期"] = pd.to_datetime(table["预计交期"]).dt.strftime("%Y-%m-%d %H:%M").astype(str)
    table["生产开始时间"] = table["生产开始时间"].dt.strftime("%Y-%m-%d %H:%M").astype(str)
    table["生产结束时间"] = table["生产结束时间"].dt.strftime("%Y-%m-%d %H:%M").astype(str)
for table in [project_delivery_df]:
    table["项目交付时间"] = table["项目交付时间"].dt.strftime("%Y-%m-%d %H:%M").astype(str)

# **📌 输出前加回前缀**
df["材料材质"] = df.apply(lambda row: restore_material(row["原始材料材质"], row["材料材质"]), axis=1)

# **📌 删除临时列**
df = df.drop(columns=["原始材料材质"])

# **📌 美化表格输出
def auto_adjust_excel(file_path):
    """自动调整 Excel 列宽，并设置居中对齐、表头加粗、单元格边框，逾期交付填充红色"""
    wb = load_workbook(file_path)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # 红色背景填充

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # 获取最大列数
        max_col = ws.max_column
        header_row = 1  # 假设第一行为表头

        # **📌 获取“按时交付检查”列的索引**
        delivery_check_col = None
        for col in range(1, max_col + 1):
            if ws.cell(row=header_row, column=col).value == "按时交付检查":
                delivery_check_col = col
                break

        # **📌 遍历所有行，设置单元格格式**
        for row in ws.iter_rows():
            for cell in row[:max_col]:  # 确保所有列都有边框
                cell.alignment = Alignment(horizontal="center", vertical="center")  # 居中
                cell.border = thin_border  # 添加边框

            # **📌 如果“按时交付检查”为“逾期交付”，填充红色**
            if delivery_check_col:
                check_cell = row[delivery_check_col - 1]  # openpyxl 列索引是从 0 开始
                if check_cell.value == "逾期交付":
                    check_cell.fill = red_fill  # 设置红色背景

        # **📌 计算最适合的列宽**
        column_widths = {}
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # 获取列的字母（如 A, B, C）

            for cell in col:
                if cell.value:
                    try:
                        # 计算最大字符数（中文字符算 2 个单位）
                        text_length = sum(2 if ord(c) > 255 else 1 for c in str(cell.value))
                        max_length = max(max_length, text_length)
                    except:
                        pass  # 忽略错误

            column_widths[col_letter] = max_length

        # **📌 应用计算后的列宽**
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width + 2  # 适配 Excel 的字体宽度

        # **📌 设置表头加粗**
        for cell in ws[header_row]:
            cell.font = Font(bold=True)

    wb.save(file_path)
    print(f"📊 Excel 格式优化完成: {file_path}")

# **📌 生成 Excel**
with pd.ExcelWriter(output_file) as writer:
    df[df["设备"] == "直管机"].drop(columns=["设备","是否有交期","交期排序","组1","组2","是否拆分","组最早交期","异型管机1不可生产", "项目交付时间"]).to_excel(writer, sheet_name="直管机", index=False)
    df[df["设备"] == "异型管机1"].drop(columns=["设备","交期排序","是否有交期","交期排序","组1","组2","是否拆分","组最早交期","异型管机1不可生产", "项目交付时间"]).to_excel(writer, sheet_name="异型管机1", index=False)
    df[df["设备"] == "异型管机2"].drop(columns=[ "设备","交期排序","是否有交期","交期排序","组1","组2","是否拆分","组最早交期","异型管机1不可生产", "项目交付时间"]).to_excel(writer, sheet_name="异型管机2", index=False)
    project_delivery_df.to_excel(writer, sheet_name="项目交付时间", index=False)
    df_other.to_excel(writer, sheet_name="其他", index=False)  # ✅ 正确添加‘其他’表单

# **📌 美化 Excel**
auto_adjust_excel(output_file)
print(f"✅ 排产已完成，结果保存至 {output_file}")